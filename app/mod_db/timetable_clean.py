import openpyxl
from openpyxl.utils import get_column_letter
# no tests
def timetable_clean(xlsx, outf):
    """ Takes a raw timetable XLSX file, formats and cleans it and saves it to CSV."""
    # Opens the outfile and writes the column names.
    fout = open(outf, "w")
    headings = "time,room,module,size\n"
    fout.write(headings)

    # Declares a list to hold each row of the outfile.
    outlist = []

    # Writes the times to the list of rows.
    days = ["Mon Nov 02 ", "Tue Nov 03 ", "Wed Nov 04 ", "Thu Nov 05 ", "Fri Nov 06 ",
            "Mon Nov 09 ", "Tue Nov 10 ", "Wed Nov 11 ", "Thu Nov 12 ", "Fri Nov 13 "]
    hours = ["09:00:00", "10:00:00", "11:00:00", "12:00:00", "13:00:00",
             "14:00:00", "15:00:00", "16:00:00", "17:00:00"]
    times = []
    for x in range(0, 12):
        for i in range(0, len(days)):
            for j in range(0, len(hours)):
                times.append(days[i] + hours[j])

    print(len(times))

    for i in range(0, 270):
        outlist.append(times[i] + ",")

    # Reads the infile.
    wb = openpyxl.load_workbook(xlsx)

    # Counter for outlist row and for writing sheet names.
    listno, j = 0, 0

    # Reads each sheet.
    j = -1
    for sheet in wb.worksheets:
        # Ignores the "All" sheet.
        if sheet.title == "All":
            break

        # Adds the name of the sheet to the outlist 45 times.
        j += 1
        for j in range(j, j + 90):
            outlist[j] += sheet.title + ","

        # Reads every other column.
        for x in range(2, sheet.max_column, 2):
            # Reads every row.
            for y in range(2, sheet.max_row):
                cell = sheet.cell(row=y, column=x)
                # Ignores if cell is a header.
                if cell.style.fill.start_color.index not in [4, 5, 9]:
                    # Checks if cell is grey.
                    if cell.style.fill.start_color.index == 00000000:
                        # Checks that cell is not part of legend.
                        if cell.value != "No class timetabled":
                            # Adds "None" as the module and size to the outlist.
                            outlist[listno] += "None,None\n"
                            listno += 1

                            # Adds to the list again if cell is merged (double class).
                            if get_column_letter(x) + str(y) in sheet.merged_cells:
                                outlist[listno] += "None,None\n"
                                listno += 1

                    # Checks that cell is not blank or part of the legend.
                    elif cell.value not in [None, "Unclear if class went ahead"]:
                        # Adds the module and size (adjacent rows) to the outlist.
                        outlist[listno] += str(cell.value) + "," + str(sheet.cell(row=y, column=x + 1).value) + "\n"
                        listno += 1

                        # Adds to the list again if cell is merged (double class).
                        if get_column_letter(x) + str(y) in sheet.merged_cells:
                            outlist[listno] += str(cell.value) + "," + str(sheet.cell(row=y, column=x + 1).value) + "\n"
                            listno += 1

    # Writes each row of the outfile list to the outfile.
    for i in range(0, len(outlist)):
        fout.write(outlist[i])