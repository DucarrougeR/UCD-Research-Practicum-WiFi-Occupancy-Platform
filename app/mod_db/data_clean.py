"""
NB: Some operations are hard-coded to the original dataset, particularly with regards
to time. This can be changed at a later point.
"""

import zipfile, os, re, csv, openpyxl
from openpyxl.utils import get_column_letter

def unzip(zip_path, room_path=None):
    """ Fully unzips a zipped folder full of zip files. """
    if zipfile.is_zipfile(zip_path):
        with zipfile.ZipFile(zip_path) as m_zip:
            if not room_path:
                folder_path = zip_path[0:zip_path.find(".zip")]
            else:
                folder_path = room_path

            # Iterates through items in the zip file.
            if not os.path.exists(folder_path):
                os.makedirs(folder_path)

            zipfile.ZipFile(zip_path, "r").extractall(folder_path)

            for item in m_zip.namelist():
                if zipfile.is_zipfile(folder_path + "/" + item):
                    # Regex match to get room number.
                    room = re.match(r"Client_Count_CSCI_([A-Z]\-\d{2})", item)
                    room_dir_path = folder_path + "/" + room.groups()[0]

                    if not os.path.exists(room_dir_path):
                        os.makedirs(room_dir_path)

                    # Unzips this file.
                    unzip(folder_path + "/" + item, room_dir_path)
                    os.remove(folder_path + "/" + item)
                else:
                    print(folder_path + item + " is not a zip")

            m_zip.close()
    else:
        print(zip_path + " is not zip")


def merge_logs(path, outf):
    """ Takes a directory containing folders full of log files and merges the data therein
    into a single CSV file. """
    # Opens the outfile and writes the column names.
    fout = open(outf, "w")
    headings = "room,time,associated,authenticated\n"
    fout.write(headings)

    # Loops through every CSV file in every folder in the data directory.
    for root, dirs, files in os.walk(path):
        for name in files:
            with open((os.path.join(root, name)), encoding="ISO-8859-1") as fin:
                # Finds the string "Key" and reads the remainder of the file to the outfile.
                for line in fin:
                    if "Key" in line:
                        fout.write(fin.read())

    fout.close()

def gt_clean(gtf, gt_raw, gt_clean):
    """ Takes a ground truth CSV file, strips it of unnecessary information and formats it correctly. """
    # Opens the first outfile.
    raw = open(gt_raw, "w")

    # Opens the infile and writes the relevant rows to the first outfile.
    with open(gtf) as fin:
        for line in fin:
            if "CSI Classroom OCCUPANCY" in line:
                for i in range(0, 2):
                    fin.readline()

                for i in range(0, 3):
                    raw.write(fin.readline())

                fin.readline()

                for i in range(0, 8):
                    raw.write(fin.readline())
    raw.close()

    # Opens the second outfile and writes the column names.
    clean = open(gt_clean, "w+")
    headings = "room,capacity,time,occupancy\n"
    clean.write(headings)

    # Declares a list to hold each row of the outfile.
    outlist = []

    with open(gt_raw) as raw:
        # Finds the room numbers.
        rooms = []
        for line in raw:
            if "Room No." in line:
                # Splits the line by comma into a list.
                for i in line.split(","):
                    if i != "Room No." and i != "" and i != "\n":
                       rooms.append(i)
                break

        # Adds the room numbers to the list of rows.
        for i in rooms:
            for j in range(0, 80):
                outlist.append(i + ",")

        # Finds the capacities.
        capacities = []
        for line in raw:
            if "Time" in line:
                # Splits the line by comma into a list.
                for i in line.split(","):
                    if i.isdigit():
                        for j in range(0, 80):
                            capacities.append(i)
                break

        # Adds the capacities to the list of rows.
        for i in range(0, len(outlist)):
            outlist[i] += capacities[i] + ","

    # Writes the times to the list of rows.
    days = ["Mon Nov 02 ", "Tue Nov 03 ", "Wed Nov 04 ", "Thu Nov 05 ", "Fri Nov 06 ",
            "Mon Nov 09 ", "Tue Nov 10 ", "Wed Nov 11 ", "Thu Nov 12 ", "Fri Nov 13 "]
    hours = ["09:00:00", "10:00:00", "11:00:00", "12:00:00", "13:00:00",
            "14:00:00", "15:00:00", "16:00:00"]
    times = []
    for x in range(0, 6):
        for i in range(0, len(days)):
            for j in range(0, len(hours)):
                times.append(days[i] + hours[j])

    for i in range(0, len(outlist)):
        outlist[i] += times[i] + ","

    with open(gt_raw) as raw:
        # Reads the first outfile by column, storing each column as a list.
        reader = csv.reader(raw, skipinitialspace=True)
        v1, v2, room1, v3, room2, room3, room4, room5, room6, room7, v4, v5, v6, v7, v8 = zip(*reader)
        roomcols = [room1, room2, room3, room4, room5, room6, room7]

        # Reads all the occupancy percentages into the relevant list.
        room1occ, room2occ, room3occ, room4occ, room5occ, room6occ, room7occ = [], [], [], [], [], [], []
        roomoccs = [room1occ, room2occ, room3occ, room4occ, room5occ, room6occ, room7occ]

        for i in range(0, len(roomcols)):
            for elem in roomcols[i]:
                if re.match(".*%", elem):
                    roomoccs[i].append(elem)

    # Writes each room occupancy list to the relevant element of the list of rows.
    j = 0
    for elem in roomoccs:
        for i in range(0, len(elem)):
            outlist[j] += elem[i]
            j += 1

    # Writes each row of the outfile list to the outfile.
    for i in range(0, len(outlist)):
        clean.write(outlist[i] + "\n")

    clean.close()

def timetable_clean(xlsx, outf):
    """ Takes a raw timetable XLSX file, formats and cleans it and saves it to CSV."""
    # Opens the outfile and writes the column names.
    fout = open(outf, "w")
    headings = "time,room,module,size\n"
    fout.write(headings)

    # Declares a list to hold each row of the outfile.
    outlist = []

    # Writes the times to the list of rows.
    days = ["Mon Nov 02 ", "Tue Nov 03 ", "Wed Nov 04 ", "Thu Nov 05 ", "Fri Nov 06 ",
            "Mon Nov 09 ", "Tue Nov 10 ", "Wed Nov 11 ", "Thu Nov 12 ", "Fri Nov 13 "]
    hours = ["09:00:00", "10:00:00", "11:00:00", "12:00:00", "13:00:00",
            "14:00:00", "15:00:00", "16:00:00", "17:00:00"]
    times = []
    for x in range(0, 12):
        for i in range(0, len(days)):
            for j in range(0, len(hours)):
                times.append(days[i] + hours[j])

    print(len(times))

    for i in range(0, 270):
        outlist.append(times[i] + ",")

    # Reads the infile.
    wb = openpyxl.load_workbook(xlsx)

    # Counter for outlist row and for writing sheet names.
    listno, j = 0, 0

    # Reads each sheet.
    j = -1
    for sheet in wb.worksheets:
        # Ignores the "All" sheet.
        if sheet.title == "All":
            break

        # Adds the name of the sheet to the outlist 45 times.
        j += 1
        for j in range(j, j + 90):
            outlist[j] += sheet.title + ","

        # Reads every other column.
        for x in range(2, sheet.max_column, 2):
            # Reads every row.
            for y in range(2, sheet.max_row):
                cell = sheet.cell(row=y, column=x)
                # Ignores if cell is a header.
                if cell.style.fill.start_color.index not in [4, 5, 9]:
                    # Checks if cell is grey.
                    if cell.style.fill.start_color.index == 00000000:
                        # Checks that cell is not part of legend.
                        if cell.value != "No class timetabled":
                            # Adds "None" as the module and size to the outlist.
                            outlist[listno] += "None,None\n"
                            listno += 1
                            
                            # Adds to the list again if cell is merged (double class).
                            if get_column_letter(x) + str(y) in sheet.merged_cells:
                                outlist[listno] += "None,None\n"
                                listno += 1

                    # Checks that cell is not blank or part of the legend.
                    elif cell.value not in [None, "Unclear if class went ahead"]:
                        # Adds the module and size (adjacent rows) to the outlist.
                        outlist[listno] += str(cell.value) + "," + str(sheet.cell(row=y, column=x+1).value) + "\n"
                        listno += 1

                        # Adds to the list again if cell is merged (double class).
                        if get_column_letter(x) + str(y) in sheet.merged_cells:
                            outlist[listno] += str(cell.value) + "," + str(sheet.cell(row=y, column=x+1).value) + "\n"
                            listno += 1

    # Writes each row of the outfile list to the outfile.
    for i in range(0, len(outlist)):
        fout.write(outlist[i])    
    
unzip("data/raw/CSI WiFiLogs.zip") 
merge_logs("data/raw/CSI WiFiLogs", "data/clean/logs_clean.csv")
gt_clean("data/raw/CSI Occupancy report/CSI-Table 1.csv", "data/raw/gt_raw.csv", "data/clean/gt_clean.csv")
timetable_clean("data/raw/timetables/B0.02 B0.03 B0.04 Timetable.xlsx", "data/clean/timetable_clean.csv")
